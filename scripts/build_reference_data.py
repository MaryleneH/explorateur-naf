#!/usr/bin/env python3
"""
build_reference_data.py
Construit les fichiers de référence NAF utilisés par l'Explorateur NAF.

Sources :
- NAF 2025 : TheodoreBillotte/Job-Finder (hiérarchie complète, données issues de l'INSEE)
             datagouv/apistration (codes sous-classes officiels NAF 2025)
- NAF rév.2 : SocialGouv/codes-naf (données issues de l'INSEE)
- Correspondances 2008 ↔ 2025 : la SOURCE DE VÉRITÉ est la table officielle
  Insee « Correspondances_NAFrev2-NAF2025.xlsx » (rééditée en janvier 2026).
  Le script la cherche dans data/raw/, sinon tente de la télécharger depuis
  l'Insee. À défaut seulement, il construit une APPROXIMATION par préfixe de
  classe partagé, clairement étiquetée comme telle dans chaque ligne : cette
  approximation n'est jamais présentée comme la table officielle.

Usage :
    python scripts/build_reference_data.py
    python scripts/build_reference_data.py --correspondances chemin/vers/Correspondances_NAFrev2-NAF2025.xlsx

Le site GitHub Pages n'a pas besoin de Python : il lit les CSV statiques générés.

Le fichier data/defense_qualification.csv n'est PAS généré par ce script :
il est maintenu à la main (chaque ligne portant justification, sources,
repère et limite d'interprétation) et contrôlé par validate_reference_data.py.
"""

import csv
import io
import json
import sys
import urllib.request
from pathlib import Path

DATA_DIR = Path(__file__).parent.parent / "data"
DATA_DIR.mkdir(exist_ok=True)
RAW_DIR = DATA_DIR / "raw"
RAW_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# URL des sources (toutes issues de dépôts GitHub publics hébergeant des
# données d'origine officielle INSEE / DARES)
# ---------------------------------------------------------------------------
SOURCES = {
    "naf2025_hierarchy": (
        "https://raw.githubusercontent.com/TheodoreBillotte/"
        "Job-Finder/main/data/naf_subset.csv"
    ),
    "naf2025_subclasses": (
        "https://raw.githubusercontent.com/datagouv/"
        "apistration/main/siade/lib/referentials/files/NAF2025.csv"
    ),
    "naf2008_json": (
        "https://raw.githubusercontent.com/SocialGouv/"
        "codes-naf/master/index.json"
    ),
    # Les correspondances 2008 ↔ 2025 ne dépendent plus d'une copie tierce :
    # voir load_official_pairs() (table officielle Insee, avec approximation
    # par préfixe en secours explicite).
}

NAF2025_SOURCE_URL = "https://www.insee.fr/fr/information/8201411"
NAF2008_SOURCE_URL = "https://www.insee.fr/fr/information/2406147"
NAF_CORRESP_SOURCE_URL = "https://www.insee.fr/fr/information/8201411"

# ---------------------------------------------------------------------------
# Fiches de métadonnées Insee
#
# Chaque nomenclature a son propre segment d'URL et le code s'écrit avec un
# point, jamais avec un underscore :
#   NAF 2025   https://www.insee.fr/fr/metadonnees/naf2025/sousClasse/30.32Y
#   NAF rév.2  https://www.insee.fr/fr/metadonnees/nafr2/sousClasse/30.30Z
# Utiliser le mauvais segment renvoie une fiche d'une autre nomenclature ou
# une 404. Ne jamais construire ces URL à la main ailleurs qu'ici.
# ---------------------------------------------------------------------------

INSEE_METADATA_ROOT = "https://www.insee.fr/fr/metadonnees"

INSEE_NOMENCLATURE_SEGMENT = {
    "NAF 2025": "naf2025",
    "NAF rév.2": "nafr2",
}

INSEE_LEVEL_SEGMENT = {
    "section": "section",
    "division": "division",
    "group": "groupe",
    "class": "classe",
    "subclass": "sousClasse",
}


def insee_url(version: str, level: str, code: str) -> str:
    """Construit l'URL de la fiche Insee d'un code, pour la bonne nomenclature."""
    try:
        nomenclature = INSEE_NOMENCLATURE_SEGMENT[version]
    except KeyError:
        raise ValueError(f"Nomenclature inconnue : {version!r}")
    try:
        segment = INSEE_LEVEL_SEGMENT[level]
    except KeyError:
        raise ValueError(f"Niveau inconnu : {level!r}")
    return f"{INSEE_METADATA_ROOT}/{nomenclature}/{segment}/{code}"


# ---------------------------------------------------------------------------
# Section mapping NAF rév.2 (2008) — source officielle INSEE
# Chaque division est affectée à une section selon la publication officielle.
# ---------------------------------------------------------------------------
NAF2008_SECTION_MAP = {
    "A": {
        "label": "Agriculture, sylviculture et pêche",
        "divisions": set(["01", "02", "03"]),
    },
    "B": {
        "label": "Industries extractives",
        "divisions": set(["05", "06", "07", "08", "09"]),
    },
    "C": {
        "label": "Industrie manufacturière",
        "divisions": set(
            [
                "10", "11", "12", "13", "14", "15", "16", "17",
                "18", "19", "20", "21", "22", "23", "24", "25",
                "26", "27", "28", "29", "30", "31", "32", "33",
            ]
        ),
    },
    "D": {
        "label": (
            "Production et distribution d'électricité, de gaz, de vapeur "
            "et d'air conditionné"
        ),
        "divisions": set(["35"]),
    },
    "E": {
        "label": (
            "Production et distribution d'eau ; assainissement, gestion des "
            "déchets et dépollution"
        ),
        "divisions": set(["36", "37", "38", "39"]),
    },
    "F": {
        "label": "Construction",
        "divisions": set(["41", "42", "43"]),
    },
    "G": {
        "label": (
            "Commerce ; réparation d'automobiles et de motocycles"
        ),
        "divisions": set(["45", "46", "47"]),
    },
    "H": {
        "label": "Transports et entreposage",
        "divisions": set(["49", "50", "51", "52", "53"]),
    },
    "I": {
        "label": "Hébergement et restauration",
        "divisions": set(["55", "56"]),
    },
    "J": {
        "label": "Information et communication",
        "divisions": set(["58", "59", "60", "61", "62", "63"]),
    },
    "K": {
        "label": "Activités financières et d'assurance",
        "divisions": set(["64", "65", "66"]),
    },
    "L": {
        "label": "Activités immobilières",
        "divisions": set(["68"]),
    },
    "M": {
        "label": "Activités spécialisées, scientifiques et techniques",
        "divisions": set(["69", "70", "71", "72", "73", "74", "75"]),
    },
    "N": {
        "label": "Activités de services administratifs et de soutien",
        "divisions": set(["77", "78", "79", "80", "81", "82"]),
    },
    "O": {
        "label": (
            "Administration publique et défense ; "
            "sécurité sociale obligatoire"
        ),
        "divisions": set(["84"]),
    },
    "P": {
        "label": "Enseignement",
        "divisions": set(["85"]),
    },
    "Q": {
        "label": "Santé humaine et action sociale",
        "divisions": set(["86", "87", "88"]),
    },
    "R": {
        "label": "Arts, spectacles et activités récréatives",
        "divisions": set(["90", "91", "92", "93"]),
    },
    "S": {
        "label": "Autres activités de services",
        "divisions": set(["94", "95", "96"]),
    },
    "T": {
        "label": (
            "Activités des ménages en tant qu'employeurs ; activités "
            "indifférenciées des ménages en tant que producteurs de biens "
            "et de services pour usage propre"
        ),
        "divisions": set(["97", "98"]),
    },
    "U": {
        "label": "Activités extra-territoriales",
        "divisions": set(["99"]),
    },
}

# Division → section lookup
DIV_TO_SECTION_2008 = {}
for sec, info in NAF2008_SECTION_MAP.items():
    for div in info["divisions"]:
        DIV_TO_SECTION_2008[div] = (sec, info["label"])


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def fetch(url: str, cache_name: str) -> bytes:
    """Télécharge une URL en mettant en cache dans data/raw/."""
    cache_path = RAW_DIR / cache_name
    if cache_path.exists():
        print(f"  [cache] {cache_name}")
        return cache_path.read_bytes()

    print(f"  [fetch] {url}")
    try:
        req = urllib.request.Request(
            url,
            headers={"User-Agent": "explorateur-naf-build/1.0"},
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            data = resp.read()
        cache_path.write_bytes(data)
        print(f"  [saved] {cache_name} ({len(data)} octets)")
        return data
    except Exception as exc:
        print(f"  [ERROR] Impossible de récupérer {url} : {exc}", file=sys.stderr)
        sys.exit(1)


def decode_bom(data: bytes) -> str:
    """Décode les octets en texte en gérant le BOM UTF-8 éventuel."""
    if data.startswith(b"\xef\xbb\xbf"):
        return data[3:].decode("utf-8")
    return data.decode("utf-8")


def write_csv(path: Path, fieldnames: list, rows: list) -> None:
    """Écrit un CSV UTF-8 avec virgule comme séparateur."""
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(
            f,
            fieldnames=fieldnames,
            quoting=csv.QUOTE_MINIMAL,
            extrasaction="ignore",
        )
        writer.writeheader()
        writer.writerows(rows)
    print(f"  [write] {path.name} ({len(rows)} lignes de données)")


# ---------------------------------------------------------------------------
# Générateur NAF 2025
# ---------------------------------------------------------------------------

def build_naf2025() -> None:
    """
    Construit data/naf2025.csv à partir de la hiérarchie complète.
    Source : TheodoreBillotte/Job-Finder (hiérarchie issue de l'INSEE)
    747 sous-classes attendues.
    """
    print("\n=== NAF 2025 ===")
    raw = fetch(SOURCES["naf2025_hierarchy"], "naf2025_hierarchy.csv")
    text = decode_bom(raw)

    # Lire la hiérarchie complète
    reader = csv.DictReader(io.StringIO(text))
    hier_rows = list(reader)

    # Construire des dictionnaires par niveau
    sections = {}   # code → label
    divisions = {}  # code → label
    groups = {}     # code → label
    classes = {}    # code → label
    subclasses = {} # code → label

    for row in hier_rows:
        level = row.get("level", "").strip()
        code = row.get("code", "").strip()
        label = row.get("label_fr", "").strip()
        if not code or not label:
            continue
        if level == "section":
            sections[code] = label
        elif level == "division":
            divisions[code] = label
        elif level == "group":
            groups[code] = label
        elif level == "class":
            classes[code] = label
        elif level == "sous-classe":
            subclasses[code] = label

    # Construire un lookup division → section
    # Les codes de division sont les 2 premiers caractères des codes
    # On reconstruit la hiérarchie par analyse de la liste ordonnée
    # Section = lettre unique, division = "NN" 2 chiffres
    # On parcourt la liste dans l'ordre et on associe chaque division
    # à la section la plus récente rencontrée.
    current_section = None
    div_to_section = {}
    for row in hier_rows:
        level = row.get("level", "").strip()
        code = row.get("code", "").strip()
        if level == "section":
            current_section = code
        elif level == "division" and current_section:
            div_to_section[code] = current_section

    # Construire un lookup group → division (2 premiers chars du code)
    def div_of_group(gcode):
        return gcode[:2]

    def div_of_class(ccode):
        return ccode[:2]

    def div_of_subclass(scode):
        return scode[:2]

    def group_of_class(ccode):
        # class = "NN.XY" → group = "NN.X"
        return ccode[:4]

    def class_of_subclass(scode):
        # subclass = "NN.XXL" → class = "NN.XX"
        return scode[:5]

    # Construire les lignes plates
    fieldnames = [
        "version", "section_code", "section_label",
        "division_code", "division_label",
        "group_code", "group_label",
        "class_code", "class_label",
        "subclass_code", "subclass_label",
        "notes_include", "notes_exclude",
        "source_url", "source_reference",
    ]

    output_rows = []
    for sc_code, sc_label in subclasses.items():
        div_code = div_of_subclass(sc_code)
        sec_code = div_to_section.get(div_code, "?")
        sec_label = sections.get(sec_code, "")
        div_label = divisions.get(div_code, "")
        grp_code = group_of_class(class_of_subclass(sc_code))
        grp_label = groups.get(grp_code, "")
        cls_code = class_of_subclass(sc_code)
        cls_label = classes.get(cls_code, "")

        source_url = insee_url("NAF 2025", "subclass", sc_code)
        source_reference = f"Insee — NAF 2025 — sous-classe {sc_code}"

        output_rows.append({
            "version": "NAF 2025",
            "section_code": sec_code,
            "section_label": sec_label,
            "division_code": div_code,
            "division_label": div_label,
            "group_code": grp_code,
            "group_label": grp_label,
            "class_code": cls_code,
            "class_label": cls_label,
            "subclass_code": sc_code,
            "subclass_label": sc_label,
            "notes_include": "",
            "notes_exclude": "",
            "source_url": source_url,
            "source_reference": source_reference,
        })

    write_csv(DATA_DIR / "naf2025.csv", fieldnames, output_rows)
    return output_rows


# ---------------------------------------------------------------------------
# Générateur NAF rév.2 (2008)
# ---------------------------------------------------------------------------

def build_naf2008() -> None:
    """
    Construit data/naf2008.csv à partir des données SocialGouv/codes-naf.
    Source originelle : INSEE — NAF rév.2.
    732 sous-classes attendues.
    """
    print("\n=== NAF rév.2 (2008) ===")
    raw = fetch(SOURCES["naf2008_json"], "naf2008.json")
    data = json.loads(decode_bom(raw))

    # Indexer par niveau
    divisions = {}
    groups = {}
    classes = {}
    subclasses = {}

    for item in data:
        code = item.get("id", "").strip()
        label = item.get("label", "").strip()
        if not code or not label:
            continue
        if "." not in code:
            # Division (ex: "01") ou chapitre (ignoré si alphabétique)
            if code.isdigit() or (len(code) == 2 and code[0].isdigit()):
                divisions[code] = label
        elif len(code) == 4:  # groupe "01.1"
            groups[code] = label
        elif len(code) == 5:  # classe "01.11"
            classes[code] = label
        elif len(code) == 6:  # sous-classe "01.11Z"
            subclasses[code] = label

    def group_of_class(ccode):
        return ccode[:4]

    def class_of_subclass(scode):
        return scode[:5]

    def div_of(code):
        return code[:2]

    fieldnames = [
        "version", "section_code", "section_label",
        "division_code", "division_label",
        "group_code", "group_label",
        "class_code", "class_label",
        "subclass_code", "subclass_label",
        "notes_include", "notes_exclude",
        "source_url", "source_reference",
    ]

    output_rows = []
    for sc_code, sc_label in subclasses.items():
        div_code = div_of(sc_code)
        sec_tuple = DIV_TO_SECTION_2008.get(div_code, ("?", ""))
        sec_code, sec_label = sec_tuple
        div_label = divisions.get(div_code, "")
        grp_code = group_of_class(class_of_subclass(sc_code))
        grp_label = groups.get(grp_code, "")
        cls_code = class_of_subclass(sc_code)
        cls_label = classes.get(cls_code, "")

        source_url = insee_url("NAF rév.2", "subclass", sc_code)
        source_reference = f"Insee — NAF rév.2 — sous-classe {sc_code}"

        output_rows.append({
            "version": "NAF rév.2",
            "section_code": sec_code,
            "section_label": sec_label,
            "division_code": div_code,
            "division_label": div_label,
            "group_code": grp_code,
            "group_label": grp_label,
            "class_code": cls_code,
            "class_label": cls_label,
            "subclass_code": sc_code,
            "subclass_label": sc_label,
            "notes_include": "",
            "notes_exclude": "",
            "source_url": source_url,
            "source_reference": source_reference,
        })

    write_csv(DATA_DIR / "naf2008.csv", fieldnames, output_rows)
    return output_rows


# ---------------------------------------------------------------------------
# Générateur correspondances 2008 → 2025
# ---------------------------------------------------------------------------

def build_correspondances(rows_2008, rows_2025) -> None:
    """
    Construit data/correspondances_2008_2025.csv.

    SOURCE DE VÉRITÉ : la table de correspondance officielle Insee
    (Correspondances_NAFrev2-NAF2025.xlsx, rééditée en janvier 2026).
    Le script la cherche en local (option --correspondances, sinon
    data/raw/), puis tente de la télécharger. À défaut seulement, il
    construit une APPROXIMATION par préfixe de classe partagé, étiquetée
    comme telle dans chaque ligne.

    Le type de correspondance (1→1, 1→n, n→1, n↔n) est calculé sur les
    cardinalités des paires ; la table officielle apporte en plus son
    propre type (Unique/Multiple) et la description du contenu commun.
    """
    print("\n=== Correspondances 2008 ↔ 2025 ===")

    index_2008 = {r["subclass_code"]: r for r in rows_2008}
    index_2025 = {r["subclass_code"]: r for r in rows_2025}

    lines, official = load_official_pairs(index_2008, index_2025)

    if not official:
        print("  ! Table officielle indisponible : approximation par préfixe "
              "de classe partagé (étiquetée comme telle).")
        lines = [{"c08": a, "c25": b, "type_officiel": "", "contenu": ""}
                 for a, b in approximate_pairs(index_2008, index_2025)]

    # Cardinalités → type de correspondance
    count_targets, count_sources = {}, {}
    for line in lines:
        count_targets.setdefault(line["c08"], set()).add(line["c25"])
        count_sources.setdefault(line["c25"], set()).add(line["c08"])

    def corr_type(c2008, c2025):
        n_target = len(count_targets.get(c2008, set()))
        n_source = len(count_sources.get(c2025, set()))
        if n_target == 1 and n_source == 1:
            return "1→1"
        if n_target > 1 and n_source == 1:
            return "1→n"
        if n_target == 1 and n_source > 1:
            return "n→1"
        return "n↔n"

    source_url = "https://www.insee.fr/fr/information/8181066"
    if official:
        source_reference = OFFICIAL_TABLE_REFERENCE
    else:
        source_reference = (
            "Rapprochement par préfixe de classe (approximation du site) — "
            "à confirmer avec la table officielle Insee NAF rév.2 → NAF 2025"
        )

    fieldnames = [
        "code_2008", "libelle_2008",
        "code_2025", "libelle_2025",
        "type_correspondance", "type_officiel", "contenu_commun",
        "source_url", "source_reference",
    ]

    output_rows = []
    for line in sorted(lines, key=lambda l: (l["c08"], l["c25"])):
        r2008 = index_2008.get(line["c08"], {})
        r2025 = index_2025.get(line["c25"], {})
        output_rows.append({
            "code_2008": line["c08"],
            "libelle_2008": r2008.get("subclass_label", ""),
            "code_2025": line["c25"],
            "libelle_2025": r2025.get("subclass_label", ""),
            "type_correspondance": corr_type(line["c08"], line["c25"]),
            "type_officiel": line["type_officiel"],
            "contenu_commun": line["contenu"],
            "source_url": source_url,
            "source_reference": source_reference,
        })

    write_csv(DATA_DIR / "correspondances_2008_2025.csv", fieldnames, output_rows)
    print(f"  {'Table officielle' if official else 'Approximation'} : "
          f"{len(output_rows)} relations écrites")
    return output_rows


OFFICIAL_TABLE_URL = (
    "https://www.insee.fr/fr/statistiques/fichier/8181066/"
    "Correspondances_NAFrev2-NAF2025.xlsx"
)
OFFICIAL_TABLE_REFERENCE = (
    "INSEE — Table de correspondance officielle NAF rév.2 → NAF 2025 "
    "(Correspondances_NAFrev2-NAF2025.xlsx)"
)

# Paires produites par l'approximation par préfixe mais contredites par les
# libellés officiels (voir data/README.md) : jamais réémises en mode secours.
EXCLUDED_APPROX_PAIRS = {
    ("25.40Z", "25.40Y"), ("25.61Z", "25.61Y"), ("38.21Z", "38.21Y"),
    ("38.22Z", "38.22Y"), ("38.31Z", "38.31Y"), ("38.32Z", "38.32Y"),
}


def load_official_pairs(index_2008, index_2025):
    """Charge la table officielle Insee si disponible.

    Retourne (lignes, True) si la table officielle a été lue, sinon
    ([], False). Cherche d'abord un fichier local (--correspondances ou
    data/raw/Correspondances_NAFrev2-NAF2025.xlsx), puis tente le
    téléchargement depuis l'Insee.

    Chaque ligne est un dict {c08, c25, type_officiel, contenu} :
    la réédition de janvier 2026 porte le type officiel (Unique/Multiple)
    et la description du « contenu commun » qui passe d'un code à l'autre.
    """
    path = None
    argv = sys.argv[1:]
    if "--correspondances" in argv:
        path = Path(argv[argv.index("--correspondances") + 1])
    else:
        candidate = RAW_DIR / "Correspondances_NAFrev2-NAF2025.xlsx"
        if candidate.exists():
            path = candidate
        else:
            try:
                print(f"  Téléchargement de la table officielle : {OFFICIAL_TABLE_URL}")
                req = urllib.request.Request(
                    OFFICIAL_TABLE_URL, headers={"User-Agent": "explorateur-naf"})
                with urllib.request.urlopen(req, timeout=60) as resp:
                    candidate.write_bytes(resp.read())
                path = candidate
            except Exception as error:  # noqa: BLE001 — réseau best-effort
                print(f"  ! Téléchargement impossible ({error})")

    if path is None or not path.exists():
        return [], False

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("  ! openpyxl absent : `pip install openpyxl` pour lire la table "
              "officielle — approximation utilisée en attendant.")
        return [], False

    wb = load_workbook(path, read_only=True, data_only=True)
    lines = []
    seen = set()
    for ws in wb.worksheets:
        headers = {}
        header_row = None
        for row in ws.iter_rows(min_row=1, max_row=12):
            for cell in row:
                value = str(cell.value or "")
                # Réédition janvier 2026.
                if value.startswith("Sous-classes NAF rév. 2 – codes"):
                    header_row = cell.row
                    headers["c08"] = cell.column
                elif value.startswith("Sous-classes NAF 2025 – codes"):
                    headers["c25"] = cell.column
                elif value.startswith("Type de correspondance"):
                    headers["type"] = cell.column
                elif value.startswith("Contenu commun"):
                    headers["contenu"] = cell.column
                # Édition antérieure (colonnes NAFold/NAFnew).
                elif value.startswith("NAFold-code"):
                    header_row = cell.row
                    headers["c08"] = cell.column
                elif value.startswith("NAFnew-code"):
                    headers["c25"] = cell.column
            if header_row:
                break
        if "c08" not in headers or "c25" not in headers:
            continue
        for row in ws.iter_rows(min_row=header_row + 1):
            def cell_at(key):
                col = headers.get(key)
                if col is None or col - 1 >= len(row):
                    return ""
                return str(row[col - 1].value or "").strip()
            c2008 = cell_at("c08")
            c2025 = cell_at("c25")
            if not c2008 or not c2025:
                continue
            if c2008 not in index_2008 or c2025 not in index_2025:
                print(f"  ! Ligne ignorée (code inconnu) : {c2008} → {c2025}")
                continue
            if (c2008, c2025) in seen:
                continue
            seen.add((c2008, c2025))
            lines.append({
                "c08": c2008,
                "c25": c2025,
                "type_officiel": cell_at("type"),
                "contenu": " ".join(cell_at("contenu").split()),
            })
    if lines:
        print(f"  Table officielle lue : {len(lines)} relations ({path.name})")
        return lines, True
    return [], False


def approximate_pairs(index_2008, index_2025):
    """Approximation de secours : produit cartésien par préfixe de classe."""
    base_to_2008, base_to_2025 = {}, {}
    for code in index_2008:
        base_to_2008.setdefault(code[:5], []).append(code)
    for code in index_2025:
        base_to_2025.setdefault(code[:5], []).append(code)
    pairs = []
    for base in set(base_to_2008) | set(base_to_2025):
        for c2008 in base_to_2008.get(base, []):
            for c2025 in base_to_2025.get(base, []):
                if (c2008, c2025) in EXCLUDED_APPROX_PAIRS:
                    continue
                pairs.append((c2008, c2025))
    return pairs


# ---------------------------------------------------------------------------
# Données Défense (10 entrées bien documentées)
# ---------------------------------------------------------------------------

def build_defense() -> None:
    """data/defense_qualification.csv n'est plus généré ici.

    La couche « relation Défense & dualité » est maintenue à la main :
    chaque ligne porte justification, nature de preuve, sources, repère
    et limite d'interprétation, contrôlés par validate_reference_data.py.
    Un générateur écraserait ce travail éditorial.
    """
    print("\n=== Qualification Défense ===")
    print("  Fichier maintenu à la main (non régénéré) : "
          "data/defense_qualification.csv")


def main() -> None:
    print("Constructeur de données de référence — Explorateur NAF")
    print("=" * 55)
    print(f"Répertoire de sortie : {DATA_DIR}")
    print(f"Répertoire de cache  : {RAW_DIR}")

    if "--only-correspondances" in sys.argv[1:]:
        # Reconstruire seulement la table de correspondance, à partir des
        # nomenclatures déjà présentes dans data/ (pas d'accès réseau).
        with (DATA_DIR / "naf2025.csv").open(encoding="utf-8") as f:
            rows_2025 = list(csv.DictReader(f))
        with (DATA_DIR / "naf2008.csv").open(encoding="utf-8") as f:
            rows_2008 = list(csv.DictReader(f))
    else:
        rows_2025 = build_naf2025()
        rows_2008 = build_naf2008()
    build_correspondances(rows_2008, rows_2025)
    build_defense()

    print("\n✓ Construction terminée.")
    print("\nRésumé :")
    print(f"  NAF 2025 : {len(rows_2025)} sous-classes")
    print(f"  NAF 2008 : {len(rows_2008)} sous-classes")


if __name__ == "__main__":
    main()
